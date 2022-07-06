# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import requests
import time
import openpyxl
from urllib.parse import urljoin
from bs4 import BeautifulSoup
import sys
import colorama
from colorama import Fore, Style

colorama.init()
count_sales_j = 0
flag1st = 0  # флаг первого открытия таблицы excel для записи заголовков
books = []

def pars(file_name, i, j, name, author, sale, Availability,count_pre_order):
    name_corr = name[i].text[1:]
    href = name[i].get('href')
    n = len(author[i].text)
    author_ok = author[i].text[1:n - 1]
    if len(author_ok) == 0:
        author_ok = '-'
    #with open("bookvoed.txt", "a") as file:
    #    file.write(name_corr + '\n' + href + '\n' + author_ok + '\n')
    if i < len(sale):
        m = len(sale[i].text)
        sal = sale[i].text[1:m - 3]
        sal = int(sal.replace(" ", ""))

        #with open("bookvoed.txt", "a") as file:
        #    file.write(str(sal) + ' rub\n\n')
        if i < count_pre_order:
            avabil = 'Pre order!'
        else:
            avabil = '1'
    else:
        with open("bookvoed.txt", "a") as file:
            file.write((Availability[j].text) + '\n\n')
        avabil = Availability[j].text
        j += 1
        sal = '-'
    new_book = {
        'name': name_corr,
        'author': author_ok,
        'price': sal,
        'publishing': deep_parsing(href),
        'link': href,
        'avabil': avabil,
        'age': deep_parsing_age(href),
    }
    books.append(new_book)
    console_output(i, books)
    save_to_csv(file_name, i, books)
    return j


def start():
    user_request = input('Enter your request ')
    print('I display results on request ' + user_request)
    url = correct_url(user_request)

    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'lxml')
    name = soup.find_all('a', class_='TUb os')
    sale = soup.find_all('div', class_='Ag')
    author = soup.find_all('div', class_='ps')
    Availability = soup.find_all('div', class_='Bg')
    city = soup.find_all('a', class_='qx')
    pre_order = soup.find_all('div', class_='xs')
    count_pre_order = len(pre_order)
    print(city[0].text)
    all_count = len(name)
    if all_count == 0:
        sys.exit("No results were found for your request!")
    print('Total available results for your query: ', all_count)
    if all_count == 60:
        print('\n')
        print(Fore.RED + 'Warning! There may be more results, please refine your query!')
        correct = 0
        while correct != 1:
            quest = input('Would you like to refine your request? (y/n) ')
            if quest == 'y' or quest == 'n':
                correct = 1
            else:
                print('Warning! Incorrect!')
        print(Style.RESET_ALL)
        if quest == 'y':
            start()
    correct = 0
    while correct !=1:
        count = int(input('Enter the number of results to be displayed (1-60) '))
        if count >=1 and count<=60:
            correct = 1
        else:
            print(Fore.RED + 'Warning! Incorrect!')
            print(Style.RESET_ALL)


    if count > all_count:
        print('\n')
        print(Fore.RED + 'Warning! The number entered is greater than the total number of results, all possible results will be displayed!')
        print(Style.RESET_ALL)
        count = all_count
    j = 0
    downloader(count)
    file_name = 'Search results for the query - ' + user_request + '_' + str(count) + '.xlsx'
    for i in range(0, count):
        j = pars(file_name, i, j, name, author, sale, Availability, count_pre_order)
        flag1st = 1
    sys.exit("The program completed its work correctly")


def deep_parsing(link):
    p = requests.get(link)
    sou = BeautifulSoup(p.text, 'lxml')
    publishing_house = sou.find_all('a', class_='ho')
    publishing = publishing_house[0].text
    publishing = publishing.replace("\n", "")
    if publishing == 'Перейти к характеристикам':
        publishing = '-'
    return publishing


def deep_parsing_age(link):
    p = requests.get(link)
    sou = BeautifulSoup(p.text, 'lxml')
    age = sou.find_all('div', class_='RC')
    age = age[0].text
    age = age.replace("\n", "")
    return age


def console_output(i, books):
    print(books[i]['name'], '\n', books[i]['author'], '\n| ', books[i]['publishing'], ' |\n', books[i]['age'], '\n', books[i]['link'])
    if books[i]['avabil'] != '1':
        print(books[i]['avabil'])
    else:
        print('-- ', books[i]['price'], ' rub')
    print('*' * 50)


def correct_url(request):
    base = 'https://www.bookvoed.ru/'
    url = 'books?q=' + request
    url = urljoin(base, url, allow_fragments=True)
    url = convert(url)
    print(url + '\n')
    return url


def save_to_csv(file, num, books):
    # Search results, Product name, Author, Price, Link, Availability
    # получаем лист, с которым будем работать
    sheet = wb['Search results']
    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 40
    sheet.column_dimensions['E'].width = 50
    sheet.column_dimensions['G'].width = 16
    if flag1st == 0:
        sheet.append(['Product name', 'Author', 'Publishing house', '!Age limit!', 'Link', 'Price', 'Availability'])
        sheet.freeze_panes = 'A2'

    # sheet.append([name, author, link, price])
    cell = sheet.cell(row=num + 2, column=1)
    cell.value = books[num]['name']
    cell = sheet.cell(row=num + 2, column=2)
    cell.value = books[num]['author']
    cell = sheet.cell(row=num + 2, column=3)
    cell.value = books[num]['publishing']
    cell = sheet.cell(row=num + 2, column=4)
    cell.value = books[num]['age']
    cell = sheet.cell(row=num + 2, column=5)
    cell.value = books[num]['link']
    cell = sheet.cell(row=num + 2, column=6)
    cell.value = books[num]['price']
    cell = sheet.cell(row=num + 2, column=7)
    if books[num]['avabil'] == '1':
        cell.value = 'In stock'
    else:
        cell.value = books[num]['avabil']
    wb.save(file)


def downloader(c):
    s = '█'
    for inc in range(101):
        time.sleep(0.001 * c)
        print('\r', 'Load', inc * s, str(inc), '% ', end='')
    print('Results are ready\n')



def convert(s):
    i = 0
    while s[i] == ' ':
        i += 1
    s = s[i:]
    i = len(s)
    while s[i - 1] == ' ':
        i -= 1
    s = s[:i]
    s_new = s[0]
    i = 1
    while i < len(s):
        if s[i] != ' ':
            s_new += s[i]
        elif s[i - 1] != ' ':
            s_new += '%20'
        i += 1
    return s_new


# создаем новый excel-файл
wb = openpyxl.Workbook()
# добавляем новый лист
wb.create_sheet(title='Search results', index=0)
start()
